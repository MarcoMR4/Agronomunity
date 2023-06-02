from django.urls import reverse
from django.shortcuts import redirect, render
from django.db.models import Q
from datetime import date
from django.conf import settings
import os
import locale
from django.db.models import Min, Max
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.contrib.auth.decorators import login_required
from .forms import UserLoginForm, AddWorker, AddCourtOrder, AddProducer, AddTransport, AddSquad, AddOrchard, AddSquadMember, AddOrder, AddClient, AddCaliber, AddQuality, AddTrip, AddIncident, AddPrice, AddFruit, AddRol, AddReport, SearchTrip
from django.contrib.auth import authenticate, logout, login
from .models import Trabajador, RolTrabajador, User, CamionTransporte, Cuadrilla, Productor, Huerta, MiembroCuadrilla, Cliente, Pedido, Calibre, Calidad, PedidoCalibreCalidad, OrdenCorte, ViajeCorte, ReporteCorte, Incidencia, PrecioAutorizado, FrutaHuerta
from django.contrib.auth.hashers import make_password

#dashboard 
@login_required
def index(request):
    ca = Cuadrilla.objects.filter(estatusCuadrilla='C_L').count()
    ci = Cuadrilla.objects.filter(estatusCuadrilla='C_O').count()
    ha = Huerta.objects.filter(estatusHuerta="H_D").count()
    hi = Huerta.objects.filter(estatusHuerta="H_S").count()
    pi = Pedido.objects.filter(estatusPedido="P_I").count() 
    po = Pedido.objects.filter(estatusPedido="P_O").count()
    pv = Pedido.objects.filter(estatusPedido="P_V").count()
    pc = Pedido.objects.filter(estatusPedido="P_C").count()

    return render(request, 'index.html', {
        "ca": ca, "ci": ci,
        "ha": ha, "hi": hi,
        "pi": pi, "po": po,
        "pv": pv, "pc": pc,
    })

@login_required
def help(request):
    return render(request, 'help.html')

def li(request):
    if request.method == 'GET':
        return render(request, "login.html",{
            "form": UserLoginForm
        })
    else:
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, "login.html",{
                "form": UserLoginForm,
                "error": "Usuario o contraseña incorrectos."
            })
        else:
            login(request, user)
            return redirect('index')

def lo(request):
    logout(request)
    return redirect('login')

#Encargado de bitacora
#Trabajador
@login_required
def worker(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        #Consultas necesarias para mostrar en plantilla
        form = AddWorker(request.POST)
        trabajadores = Trabajador.objects.all()
        ntrabajadores = Trabajador.objects.count()
        roles = RolTrabajador.objects.all().order_by('nombreRol')
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Trabajador'] = request.POST['Trabajador']
                    url = reverse('wd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('w')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                form = AddWorker(request.POST)
                if form.is_valid():
                    try:
                        request.session['Nombres'] = request.POST['Nombre']
                        request.session['Apellidos'] = request.POST['AP']+' '+request.POST['AM']
                        request.session['Telefono'] = request.POST['Telefono']
                        request.session['Correo'] = request.POST['Correo']
                        request.session['Rol'] = request.POST['Rol']
                        url = reverse('wr')
                        return redirect(url)
                    except Exception as e:
                        request.session['Operacion'] = 0
                        request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                        url = reverse('w')
                        return redirect(url)
                else:
                    request.session['Error'] = "Datos incorrectos para el registro, intente nuevamente."
                    return render(request, "user_enc_bit/worker.html", {
                        'form':form,
                        'trabajadores': trabajadores,
                        'ntrabajadores': ntrabajadores,
                        'roles' : roles,
                        "error": request.session['Error']
                    })
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Trabajador'] = request.POST['Trabajador']
                    request.session['Nusuario'] = request.POST['Nusuario']
                    request.session['Nombres'] = request.POST['Nombre']
                    request.session['Apellidos'] = request.POST['Apellidos']
                    request.session['Telefono'] = request.POST['Telefono']
                    request.session['Correo'] = request.POST['Correo']
                    request.session['Rol'] = request.POST['Rol']
                    url = reverse('wm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('w')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/worker.html", {
                    'form':form,
                    'trabajadores': trabajadores,
                    'ntrabajadores': ntrabajadores,
                    'roles' : roles,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/worker.html", {
                    'form':form,
                    'trabajadores': trabajadores,
                    'ntrabajadores': ntrabajadores,
                    'roles' : roles,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_enc_bit/worker.html", {
                    'form':form,
                    'trabajadores': trabajadores,
                    'ntrabajadores': ntrabajadores,
                    'roles' : roles
                })
    else: 
        return render(request, 'denied.html')

@login_required
def workerDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtiene el trabajador y elimina
            trabajador = request.session.get('Trabajador')
            t = Trabajador.objects.get(id=trabajador)
            u = User.objects.get(id = t.usuario.id)
            u.delete() 
            t.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Trabajador eliminado correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('w')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def workerRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            nombres=request.session.get('Nombres')
            apellidos=request.session.get('Apellidos')
            palabras = nombres.split() + apellidos.split()
            iniciales = []
            for palabra in palabras:
                iniciales.append(palabra[0].upper())
            nusuario = ''.join(iniciales)
            usuario = User.objects.create(
                username='Agro-'+nusuario,
                first_name=request.session.get('Nombres'),
                last_name=request.session.get('Apellidos'),
                password= make_password(request.session.get('Telefono')),
                email = request.session.get('Correo')
            )

            r = RolTrabajador.objects.get(nombreRol=request.session.get('Rol'))
            Trabajador.objects.create(
                telefono=request.session.get('Telefono'),
                rol=r,
                usuario=usuario
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Trabajador registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('w')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def workerModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtienen los datos y se modifican
            trabajador = Trabajador.objects.get(id=request.session.get('Trabajador'))
            usuario = User.objects.get(id=trabajador.usuario.id)
            rol = RolTrabajador.objects.get(nombreRol=request.session.get('Rol'))
            usuario.username=request.session.get('Nusuario')
            usuario.first_name =request.session.get('Nombres')
            usuario.last_name=request.session.get('Apellidos')
            usuario.email=request.session.get('Correo')
            usuario.save()
            trabajador.telefono=request.session.get('Telefono')
            trabajador.rol=rol
            trabajador.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('w')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

#Productor
@login_required
def producer(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('I_C', 'E_A'):
        #Consultas necesarias para mostrar en plantilla
        form = AddProducer(request.POST)
        productores = Productor.objects.all()
        nproductores = Productor.objects.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Productor'] = request.POST['Productor']
                    url = reverse('pd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('p')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                form = AddProducer(request.POST)
                if form.is_valid():
                    try:
                        request.session['Nombre'] = request.POST['Nombre']
                        request.session['AP'] = request.POST['AP']
                        request.session['AM'] = request.POST['AM']
                        request.session['Telefono'] = request.POST['Telefono']
                        url = reverse('pr')
                        return redirect(url)
                    except Exception as e:
                        request.session['Operacion'] = 0
                        request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                        url = reverse('p')
                        return redirect(url)
                else:
                    request.session['Error'] = "Datos incorrectos para el registro, intente nuevamente."
                    return render(request, 'user_ing_campo/producer.html', {
                        'form':form,
                        "error": request.session['Error'],
                        'productores': productores,
                        'nproductores': nproductores
                    })
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Productor'] = request.POST['Productor']
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['Telefono'] = request.POST['Telefono']
                    url = reverse('pm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('p')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_ing_campo/producer.html', {
                    'form':form,
                    "mensaje": request.session['Mensaje'],
                    'productores': productores,
                    'nproductores': nproductores
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_ing_campo/producer.html', {
                    'form':form,
                    "error": request.session['Error'],
                    'productores': productores,
                    'nproductores': nproductores
                })
            else:
                return render(request, "user_ing_campo/producer.html", {
                    'form':form,
                    'productores': productores,
                    'nproductores': nproductores
                })
    else: 
        return render(request, 'denied.html')

@login_required
def producerRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('I_C', 'E_A'):
        try:

            Productor.objects.create(
                nombre=request.session.get('Nombre'),
                apellidoP=request.session.get('AP'),
                apellidoM=request.session.get('AM'),
                telefono=request.session.get('Telefono'),
            )

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Productor registrado correctamente."

        except Exception as e:
            
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."

        url = reverse('p')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def producerDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('I_C', 'E_A'):
        try:
            #Se obtiene el trabajador y elimina
            productor = request.session.get('Productor')
            p = Productor.objects.get(id=productor)
            p.delete()
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Productor eliminado correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('p')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def producerModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('I_C', 'E_A'):
        try:
            #Se obtienen los datos y se modifican
            productor = Productor.objects.get(id=request.session.get('Productor'))
            productor.nombre=request.session.get('Nombre')
            productor.apellidoP=request.session.get('AP')
            productor.apellidoM=request.session.get('AM')
            productor.telefono=request.session.get('Telefono')
            productor.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('p')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

#Cuadrillas
@login_required
def squad(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        form = AddSquad(request.POST)
        cuadrillas = Cuadrilla.objects.all()
        ncuadrillas = Cuadrilla.objects.all().count()
        numJefGer = Trabajador.objects.filter(Q(rol__nomenclaturaRol__exact='G_C') | Q(rol__nomenclaturaRol__exact='J_C')).count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    url = reverse('sd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('s')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                form = AddSquad(request.POST)
                if form.is_valid():
                    try:
                        request.session['Nombre'] = request.POST['Nombre']
                        request.session['Gerente'] = request.POST['Gerente']
                        request.session['Jefe'] = request.POST['Jefe']
                        request.session['Ubicacion'] = request.POST['Ubicacion']
                        request.session['Estatus'] = request.POST['Estatus']
                        url = reverse('sr')
                        return redirect(url)
                    except Exception as e:
                        request.session['Operacion'] = 0
                        request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                        url = reverse('s')
                        return redirect(url)
                else:
                    request.session['Error'] = "Datos incorrectos para el registro, intente nuevamente."
                    return render(request, 'user_enc_bit/squad.html', {
                        'form':form,
                        'numJefGer':numJefGer,
                        'ncuadrillas':ncuadrillas,
                        "error": request.session['Error'],
                        'cuadrillas': cuadrillas
                    })
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    url = reverse('sm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('s')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/squad.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    "mensaje": request.session['Mensaje'],
                    'cuadrillas': cuadrillas
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/squad.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    "error": request.session['Error'],
                    'cuadrillas': cuadrillas
                })
            else:
                return render(request, 'user_enc_bit/squad.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    'cuadrillas': cuadrillas
                })
    else: 
        return render(request, 'denied.html')

@login_required
def squadDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            #Se obtiene la cuadrilla y elimina
            cuadrilla = request.session.get('Cuadrilla')
            c = Cuadrilla.objects.get(id=cuadrilla)
            c.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cuadrilla eliminada correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('s')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            #Se obtienen los datos y se crea la cuadrilla
            gerente=Trabajador.objects.get(id=request.session.get('Gerente'))
            jefe=Trabajador.objects.get(id=request.session.get('Jefe'))
            Cuadrilla.objects.create(
                nombreCuadrilla=request.session.get('Nombre'),
                ubicacionCuadrilla=request.session.get('Ubicacion'),
                estatusCuadrilla=request.session.get('Estatus'),
                idJefeCuadrilla= jefe,
                idGerenteCuadrilla=gerente
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cuadrilla registrada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('s')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
        formm = AddSquadMember()
        miembros = MiembroCuadrilla.objects.filter(idCuadrilla=cuadrilla.id)
        trabajadores = Trabajador.objects.filter(Q(rol__nomenclaturaRol__exact='G_C') | Q(rol__nomenclaturaRol__exact='J_C'))
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Miembro'] = request.POST['Miembro']
                    url = reverse('smd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('sm')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['noImss'] = request.POST['noImss']
                    url = reverse('smr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('sm')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['Ubicacion'] = request.POST['Ubicacion']
                    request.session['Estatus'] = request.POST['Estatus']
                    request.session['Gerente'] = request.POST['Gerente']
                    request.session['Jefe'] = request.POST['Jefe']
                    url = reverse('sms')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('sm')
                    return redirect(url)
            elif request.POST['Id']=='editar':
                try:
                    request.session['Miembro'] = request.POST['Miembro']
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['noImss'] = request.POST['noImss']
                    url = reverse('smm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo guardar los datos, intente de nuevo."
                    url = reverse('sm')
                    return redirect(url)
        else:

            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/squadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    "mensaje": request.session['Mensaje'],
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/squadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    "error": request.session['Error'],
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
            else:
                return render(request, 'user_enc_bit/squadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
    else: 
        return render(request, 'denied.html')

@login_required
def squadMemberSave(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            #Se obtienen los datos y se modifican
            gerente=Trabajador.objects.get(id=request.session.get('Gerente'))
            jefe=Trabajador.objects.get(id=request.session.get('Jefe'))

            cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
            cuadrilla.nombreCuadrilla=request.session.get('Nombre')
            cuadrilla.estatusCuadrilla=request.session.get('Estatus')
            cuadrilla.ubicacionCuadrilla=request.session.get('Ubicacion')
            cuadrilla.idJefeCuadrilla=jefe
            cuadrilla.idGerenteCuadrilla=gerente
            cuadrilla.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron los datos correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('sm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadMemberDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            #Se obtiene el miembro y elimina
            miembro = request.session.get('Miembro')
            m = MiembroCuadrilla.objects.get(id=miembro)
            m.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Miembro eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('sm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadMemberModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            #Se obtienen los datos y se modifican
            miembro = MiembroCuadrilla.objects.get(id=request.session.get('Miembro'))
            miembro.nombre=request.session.get('Nombre')
            miembro.apellidoP=request.session.get('AP')
            miembro.apellidoM=request.session.get('AM')
            miembro.noImss = request.session.get('noImss')
            miembro.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('sm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadMemberRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'G_C'):
        try:
            cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
            #Se obtienen los datos y se crea el miembro
            MiembroCuadrilla.objects.create(
                nombre=request.session.get('Nombre'),
                apellidoP=request.session.get('AP'),
                apellidoM=request.session.get('AM'),
                noImss = request.session.get('noImss'),
                idCuadrilla=cuadrilla
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Miembro registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('sm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

#Encargado de transporte
@login_required
def transport(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_T'):
        form = AddTransport(request.POST)
        camiones = CamionTransporte.objects.all()
        ncamiones = camiones.count()
        choferes = Trabajador.objects.filter(rol__nomenclaturaRol__exact='C_T')
        nchoferes = choferes.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Camion'] = request.POST['Camion']
                    url = reverse('td')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('t')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                form = AddTransport(request.POST) 
                if form.is_valid():
                    try:
                        request.session['Chofer'] = request.POST['Chofer']
                        request.session['Placa'] = request.POST['Placa']
                        request.session['Modelo'] = request.POST['Modelo']
                        request.session['Capacidad'] = request.POST['Capacidad']
                        request.session['Tipo'] = request.POST['Tipo']
                        request.session['Descripcion'] = request.POST['Descripcion']
                        request.session['Candado'] = request.POST['Candado']
                        request.session['Estatus'] = request.POST['Estatus']
                        url = reverse('tr')
                        return redirect(url)
                    except Exception as e:
                        request.session['Operacion'] = 0
                        request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                        url = reverse('t')
                        return redirect(url)
                else:
                    request.session['Error'] = "Datos incorrectos para el registro, intente nuevamente."
                    return render(request, 'user_enc_trans/transport.html', {
                        'form':form,
                        "error": request.session['Error'],
                        'choferes': choferes,
                        'camiones': camiones,
                        'nchoferes': nchoferes,
                        'ncamiones': ncamiones
                    })
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Camion'] = request.POST['Camion']
                    request.session['Chofer'] = request.POST['Chofer']
                    request.session['Placa'] = request.POST['Placa']
                    request.session['Modelo'] = request.POST['Modelo']
                    request.session['Capacidad'] = request.POST['Capacidad']
                    request.session['Tipo'] = request.POST['Tipo']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    request.session['Candado'] = request.POST['Candado']
                    request.session['Estatus'] = request.POST['Estatus']
                    url = reverse('tm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('t')
                    return redirect(url)
        else:

            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_trans/transport.html', {
                    'form':form,
                    "mensaje": request.session['Mensaje'],
                    'choferes': choferes,
                    'camiones': camiones,
                    'nchoferes': nchoferes,
                    'ncamiones': ncamiones
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_trans/transport.html', {
                    'form':form,
                    "error": request.session['Error'],
                    'choferes': choferes,
                    'camiones': camiones,
                    'nchoferes': nchoferes,
                    'ncamiones': ncamiones
                })
            else:
                return render(request, 'user_enc_trans/transport.html', {
                    'form':form,
                    'choferes': choferes,
                    'camiones': camiones,
                    'nchoferes': nchoferes,
                    'ncamiones': ncamiones
                })
    else: 
        return render(request, 'denied.html')

@login_required
def transportDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_T'):
        try:
            #Se obtiene el miembro y elimina
            camion = request.session.get('Camion')
            c = CamionTransporte.objects.get(id=camion)
            c.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Vehiculo eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('t')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def transportRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_T'):
        try:
            chofer = Trabajador.objects.get(id=request.session.get('Chofer'))
            #Se obtienen los datos y se crea el miembro
            CamionTransporte.objects.create(
                capacidadTransporte=request.session.get('Capacidad'),
                placaTransporte=request.session.get('Placa'),
                tipoTransporte=request.session.get('Tipo'),
                descripcionTransporte=request.session.get('Descripcion'),
                modeloTransporte=request.session.get('Modelo'),
                candadoTransporte=request.session.get('Candado'),
                estatusTransporte=request.session.get('Estatus'),
                idChoferTransporte=chofer
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Vehiculo registrado correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('t')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def transportModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_T'):
        try:
            #Se obtienen los datos y se modifican
            chofer = Trabajador.objects.get(id=request.session.get('Chofer'))
            camion = CamionTransporte.objects.get(id=request.session.get('Camion'))
            camion.idChoferTransporte=chofer
            camion.placaTransporte =request.session.get('Placa')
            camion.modeloTransporte =request.session.get('Modelo')
            camion.tipoTransporte =request.session.get('Tipo')
            camion.descripcionTransporte =request.session.get('Descripcion')
            camion.candadoTransporte =request.session.get('Candado')
            camion.capacidadTransporte =request.session.get('Capacidad')
            camion.estatusTransporte = request.session.get('Estatus')
            camion.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('t')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

def orchard(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'I_C', 'E_A'):

        form = AddOrchard(request.POST)
        huertas = Huerta.objects.all()
        nhuertas = huertas.count()
        productores = Productor.objects.all().order_by('apellidoP')
        nproductores = productores.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Huerta'] = request.POST['Huerta']
                    url = reverse('od')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('o')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                form = AddOrchard(request.POST)
                if form.is_valid():
                    try:
                        request.session['Nombre'] = request.POST['Nombre']
                        request.session['Fruta'] = request.POST['Fruta']
                        request.session['Ubicacion'] = request.POST['Ubicacion']
                        request.session['Localizacion'] = request.POST['Localizacion']
                        request.session['Clave'] = request.POST['Clave']
                        request.session['Inocuidad'] = request.POST['Inocuidad']
                        request.session['Productor'] = request.POST['Productor']
                        request.session['Estatus'] = request.POST['Estatus']
                        url = reverse('or')
                        return redirect(url)
                    except Exception as e:
                        request.session['Operacion'] = 0
                        request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                        url = reverse('o')
                        return redirect(url)
                else:
                    request.session['Error'] = "Datos incorrectos para el registro, intente nuevamente."
                    return render(request, 'user_enc_bit/orchard.html', {
                        'form':form,
                        "error": request.session['Error'],
                        'huertas': huertas,
                        'nhuertas': nhuertas,
                        'nproductores': nproductores,
                        'productores': productores
                    })
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Huerta'] = request.POST['Huerta']
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['Fruta'] = request.POST['Fruta']
                    request.session['Ubicacion'] = request.POST['Ubicacion']
                    request.session['Localizacion'] = request.POST['Localizacion']
                    request.session['Clave'] = request.POST['Clave']
                    request.session['Inocuidad'] = request.POST['Inocuidad']
                    request.session['Productor'] = request.POST['Productor']
                    request.session['Estatus'] = request.POST['Estatus']
                    url = reverse('om')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('o')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/orchard.html', {
                    'form':form,
                    "mensaje": request.session['Mensaje'],
                    'huertas': huertas,
                    'nhuertas': nhuertas,
                    'nproductores': nproductores,
                    'productores': productores
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/orchard.html', {
                    'form':form,
                    "error": request.session['Error'],
                    'huertas': huertas,
                    'nhuertas': nhuertas,
                    'nproductores': nproductores,
                    'productores': productores
                })
            else:
                return render(request, 'user_enc_bit/orchard.html', {
                    'form':form,
                    'huertas': huertas,
                    'nhuertas': nhuertas,
                    'nproductores': nproductores,
                    'productores': productores
                })
    else: 
        return render(request, 'denied.html')

@login_required
def orchardDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T', 'I_C', 'E_A'):
        try:
            #Se obtiene el miembro y elimina
            huerta = request.session.get('Huerta')
            h = Huerta.objects.get(id=huerta)
            h.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Huerta eliminada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('o')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def orchardRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T', 'I_C', 'E_A'):
        try:
            #Se obtienen los datos y se crea el miembro
            productor = Productor.objects.get(nombre=request.session.get('Productor'))
            Huerta.objects.create(
                nombreHuerta=request.session.get('Nombre'),
                frutaHuerta=request.session.get('Fruta'),
                ubicacionHuerta=request.session.get('Ubicacion'),
                localizacionHuerta=request.session.get('Localizacion'),
                claveSagarpaHuerta=request.session.get('Clave'),
                estatusInocuidadHuerta=request.session.get('Inocuidad'),
                estatusHuerta=request.session.get('Estatus'),
                idProductor=productor
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Huerta registrada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('o')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def orchardModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T', 'I_C', 'E_A'):
        try:
            #Se obtienen los datos y se modifican
            productor=Productor.objects.get(id=request.session.get('Productor'))

            huerta = Huerta.objects.get(id=request.session.get('Huerta'))
            huerta.nombreHuerta=request.session.get('Nombre')
            huerta.frutaHuerta=request.session.get('Fruta')
            huerta.ubicacionHuerta=request.session.get('Ubicacion')
            huerta.localizacionHuerta=request.session.get('Localizacion')
            huerta.claveSagarpaHuerta=request.session.get('Clave')
            huerta.estatusInocuidadHuerta=request.session.get('Inocuidad')
            huerta.estatusHuerta=request.session.get('Estatus')
            huerta.idProductor=productor
            huerta.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('o')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

def trip(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T'):
        form = AddTrip()
        vc = ViajeCorte.objects.all().order_by('-fechaViaje')
        nvc = vc.count()
        nop = OrdenCorte.objects.filter(idPedido__estatusPedido='P_O').count()
        pcc = PedidoCalibreCalidad.objects.all()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Viaje'] = request.POST['Viaje']
                    url = reverse('ctd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('ct')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Camion1'] = request.POST['Camion1']
                    request.session['Camion2'] = request.POST['Camion2']
                    request.session['Salida'] = request.POST['Salida']
                    request.session['Orden'] = request.POST['Orden']
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    request.session['Punto'] = request.POST['Punto']
                    url = reverse('ctr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('ct')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Viaje'] = request.POST['Viaje']
                    request.session['Camion1'] = request.POST['Camion1']
                    request.session['Camion2'] = request.POST['Camion2']
                    request.session['Salida'] = request.POST['Salida']
                    request.session['Orden'] = request.POST['Orden']
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    request.session['Punto'] = request.POST['Punto']
                    url = reverse('ctm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('ct')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/trip.html', {
                    'form':form,
                    "mensaje": request.session['Mensaje'],
                    'nop': nop,
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/trip.html', {
                    'form':form,
                    "error": request.session['Error'],
                    'nop': nop,
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
            else:
                return render(request, 'user_enc_bit/trip.html', {
                    'form':form,
                    'nop': nop,
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
    else: 
        return render(request, 'denied.html')

@login_required
def tripDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T'):
        try:
            #Se obtiene el miembro y elimina
            v = ViajeCorte.objects.get(id=request.session.get('Viaje'))
            cuadrilla = Cuadrilla.objects.get(id=v.idCuadrilla.id)
            cuadrilla.estatusCuadrilla = 'C_L'
            cuadrilla.save()
            o = OrdenCorte.objects.get(id=v.idOrdenCorte.id)
            p = Pedido.objects.get(id=o.idPedido.id)
            p.estatusPedido = 'P_O'
            p.save()
            cuadrilla.save()
            v.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Viaje eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('ct')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def tripRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'E_T'):
        try:
            #Se obtienen los datos y se crea el miembro
            camion1 = CamionTransporte.objects.get(id=request.session.get('Camion1'))
            camion2 = CamionTransporte.objects.get(id=request.session.get('Camion2'))
            orden = OrdenCorte.objects.get(id=request.session.get('Orden'))
            pedido = Pedido.objects.get(id=orden.idPedido.id)
            cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
            cuadrilla.estatusCuadrilla = 'C_O'
            ViajeCorte.objects.create(
                fechaViaje=date.today(),
                idCamionTransporte=camion1,
                idCamionSecundarioTransporte=camion2,
                horaSalida=request.session.get('Salida'),
                idOrdenCorte=orden,
                idCuadrilla=cuadrilla,
                puntoReunion=request.session.get('Punto')
            )
            pedido.estatusPedido = 'P_V'
            cuadrilla.save()
            pedido.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Viaje registrado correctamente."
        except Exception as e:
            print(e)
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('ct')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def order(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        form = AddOrder()
        pedidos = Pedido.objects.all()
        npedidos = pedidos.count()
        nclientes = Cliente.objects.all().count()
        pcc = PedidoCalibreCalidad.objects.all()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Pedido'] = request.POST['Pedido']
                    url = reverse('sod')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('so')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Numero'] = request.POST['Numero']
                    request.session['Kilos'] = request.POST['Tkilos']
                    request.session['Mercado'] = request.POST['Mercado']
                    request.session['Destino'] = request.POST['Destino']
                    request.session['Cliente'] = request.POST['Cliente']
                    request.session['Observacion'] = request.POST['Observacion']
                    request.session['DatosT'] = request.POST['DatosT']
                    url = reverse('sor')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('so')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Pedido'] = request.POST['Pedido']
                    request.session['Kilos'] = request.POST['Kilos']
                    request.session['Mercado'] = request.POST['Mercado']
                    request.session['Destino'] = request.POST['Destino']
                    url = reverse('som')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('so')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/order.html', {
                    "mensaje": request.session['Mensaje'],
                    'form':form,
                    'pedidos': pedidos,
                    'pcc': pcc,
                    'npedidos': npedidos,
                    'nclientes': nclientes
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/order.html', {
                    "error": request.session['Error'],
                    'form':form,
                    'pcc': pcc,
                    'pedidos': pedidos,
                    'npedidos': npedidos,
                    'nclientes': nclientes
                })
            else:
                return render(request, 'user_enc_ventas/order.html', {
                    'form':form,
                    'pcc': pcc,
                    'pedidos': pedidos,
                    'npedidos': npedidos,
                    'nclientes': nclientes
                })
    else: 
        return render(request, 'denied.html')

@login_required
def orderDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtiene el miembro y elimina
            p = Pedido.objects.get(id=request.session.get('Pedido'))
            p.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Pedido eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('so')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def orderRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se crea el miembro
            cliente = Cliente.objects.get(id=request.session.get('Cliente'))
            kilos = float(request.session.get('Kilos'))
            pallets = int(kilos // 1000)
            datos = request.session.get('DatosT')
            # Eliminar la coma al final de la cadena
            datos = datos.rstrip(',')
            # Dividir la cadena en segmentos de tres elementos
            segmentos = datos.split(',')
            # Iterar sobre los segmentos
            p = Pedido.objects.create(
                idTrabajador=request.user.trabajador,
                idCliente = cliente,
                numeroPedido=request.session.get('Numero'),
                fechaPedido= date.today(),
                totalKilosPedido=kilos,
                totalPalletsPedido = pallets, 
                mercadoPedido=request.session.get('Mercado'),
                observacionPedido=request.session.get('Observacion'),
                destinoPedido=request.session.get('Destino'),
                estatusPedido='P_I'
            )

            for i in range(0, len(segmentos), 3):
                # Obtener los datos individuales
                D1 = segmentos[i]
                D2 = segmentos[i+1]
                D3 = segmentos[i+2]

                # Realizar las consultas
                calidad = Calidad.objects.get(id=D1)
                calibre = Calibre.objects.get(id=D2)

                # Realizar la inserción
                PedidoCalibreCalidad.objects.create(
                    idPedido=p,
                    idCalibre=calibre,
                    idCalidad=calidad,
                    cantidadCC=D3,
                )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Pedido registrado correctamente."
        except Exception as e:
            print(e)
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
        url = reverse('so')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def orderModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se modifican
            kilos = float(request.session.get('Kilos'))
            pallets = int(kilos // 1000)

            pedido = Pedido.objects.get(id=request.session.get('Pedido'))
            pedido.mercadoPedido = request.session.get('Mercado')
            pedido.destinoPedido = request.session.get('Destino')
            pedido.totalKilosPedido = kilos
            pedido.totalPalletsPedido = pallets
            pedido.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('so')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def client(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        form = AddClient()
        clientes = Cliente.objects.all()
        nclientes = clientes.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Cliente'] = request.POST['Cliente']
                    url = reverse('cd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('c')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['RFC'] = request.POST['RFC']
                    url = reverse('cr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('c')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Cliente'] = request.POST['Cliente']
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['RFC'] = request.POST['RFC']
                    url = reverse('cm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('c')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/client.html', {
                    "mensaje": request.session['Mensaje'],
                    'form':form,
                    'clientes': clientes,
                    'nclientes': nclientes
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/client.html', {
                    "error": request.session['Error'],
                    'form':form,
                    'clientes': clientes,
                    'nclientes': nclientes
                })
            else:
                return render(request, 'user_enc_ventas/client.html', {
                    'form':form,
                    'clientes': clientes,
                    'nclientes': nclientes
                })
    else: 
        return render(request, 'denied.html')

@login_required
def clientDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtiene el miembro y elimina
            c = Cliente.objects.get(id=request.session.get('Cliente'))
            c.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cliente eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('c')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def clientRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se crea el miembro
            Cliente.objects.create(
                nombreCliente=request.session.get('Nombre'),
                apellidoPCliente=request.session.get('AP'),
                apellidoMCliente=request.session.get('AM'),
                rfcCliente=request.session.get('RFC')
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cliente registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('c')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def clientModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se modifican

            cliente = Cliente.objects.get(id=request.session.get('Cliente'))
            cliente.nombreCliente = request.session.get('Nombre')
            cliente.apellidoPCliente = request.session.get('AP')
            cliente.apellidoMCliente = request.session.get('AM')
            cliente.rfcCliente = request.session.get('RFC')
            cliente.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('c')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def quality(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        formcalibre = AddCaliber()
        formcalidad = AddQuality()
        calidades = Calidad.objects.all()
        ncalidades = calidades.count()
        calibres = Calibre.objects.all()
        ncalibres = calibres.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminarcalibre':
                try:
                    request.session['Calibre'] = request.POST['Calibre']
                    url = reverse('qdc')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
            if request.POST['Id']=='eliminarcalidad':
                try:
                    request.session['Calidad'] = request.POST['Calidad']
                    url = reverse('qdq')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
            elif request.POST['Id']=='agregarcalibre':
                try:
                    request.session['Numero'] = request.POST['Numero']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    url = reverse('qrc')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
            elif request.POST['Id']=='agregarcalidad':
                try:
                    request.session['Numero'] = request.POST['Numero']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    url = reverse('qrq')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
            elif request.POST['Id']=='modificarcalidad':
                try:
                    request.session['Calidad'] = request.POST['Calidad']
                    request.session['Numero'] = request.POST['Numero']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    url = reverse('qmq')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
            elif request.POST['Id']=='modificarcalibre':
                try:
                    request.session['Calibre'] = request.POST['Calibre']
                    request.session['Numero'] = request.POST['Numero']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    url = reverse('qmc')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('q')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/quality.html', {
                    "mensaje": request.session['Mensaje'],
                    'formcalidad':formcalidad,
                    'formcalibre':formcalibre,
                    'calidades': calidades,
                    'ncalidades': ncalidades,
                    'calibres': calibres,
                    'ncalibres': ncalibres
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ventas/quality.html', {
                    "error": request.session['Error'],
                    'formcalidad':formcalidad,
                    'formcalibre':formcalibre,
                    'calidades': calidades,
                    'ncalidades': ncalidades,
                    'calibres': calibres,
                    'ncalibres': ncalibres
                })
            else:
                return render(request, 'user_enc_ventas/quality.html', {
                    'formcalidad':formcalidad,
                    'formcalibre':formcalibre,
                    'calidades': calidades,
                    'ncalidades': ncalidades,
                    'calibres': calibres,
                    'ncalibres': ncalibres
                })
    else: 
        return render(request, 'denied.html')

@login_required
def qualityDeleteQuality(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtiene el miembro y elimina
            c = Calidad.objects.get(id=request.session.get('Calidad'))
            c.delete()
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Calidad eliminada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."

        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def qualityDeleteCaliber(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtiene el miembro y elimina
            c = Calibre.objects.get(id=request.session.get('Calibre'))
            c.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Calibre eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def qualityRegisterQuality(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se crea el miembro
            Calidad.objects.create(
                numCalidad=request.session.get('Numero'),
                descripcionCalidad=request.session.get('Descripcion')
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Calidad registrada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def qualityRegisterCaliber(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se crea el miembro
            Calibre.objects.create(
                numCalibre=request.session.get('Numero'),
                descripcionCalibre=request.session.get('Descripcion')
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Calibre registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def qualityModifyQuality(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se modifican

            calidad = Calidad.objects.get(id=request.session.get('Calidad'))
            calidad.numCalidad = request.session.get('Numero')
            calidad.descripcionCalidad = request.session.get('Descripcion')
            calidad.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

def qualityModifyCaliber(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_V'):
        try:
            #Se obtienen los datos y se modifican

            calibre = Calibre.objects.get(id=request.session.get('Calibre'))
            calibre.numCalibre = request.session.get('Numero')
            calibre.descripcionCalibre = request.session.get('Descripcion')
            calibre.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def courtOrder(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):

        form = AddCourtOrder()
        npi = Pedido.objects.filter(estatusPedido='P_I').count()
        oc = OrdenCorte.objects.order_by('-fechaOrden')
        pcc = PedidoCalibreCalidad.objects.all()
        noc = oc.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Orden'] = request.POST['Orden']
                    url = reverse('cod')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('co')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Huerta'] = request.POST['Huerta']
                    request.session['Corte'] = request.POST['Corte']
                    request.session['Pedido'] = request.POST['Pedido']
                    url = reverse('cor')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('co')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/courtOrder.html', {
                    "mensaje": request.session['Mensaje'],
                    'npi':npi,
                    'pcc':pcc,
                    'oc':oc,
                    'noc': noc,
                    'form': form
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_bit/courtOrder.html', {
                    "error": request.session['Error'],
                    'npi':npi,
                    'oc':oc,
                    'pcc':pcc,
                    'noc': noc,
                    'form': form
                })
            else:
                return render(request, 'user_enc_bit/courtOrder.html', {
                    'npi':npi,
                    'oc':oc,
                    'noc': noc,
                    'pcc':pcc,
                    'form': form
                })
    else: 
        return render(request, 'denied.html')

@login_required
def courtOrderDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):

        try:
            #Se obtiene el miembro y elimina
            o = OrdenCorte.objects.get(id=request.session.get('Orden'))
            pedido = Pedido.objects.get(id=o.idPedido.id)
            pedido.estatusPedido='P_I'
            pedido.save()
            o.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Orden de corte eliminada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('co')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def courtOrderRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):

        try:
            huerta = Huerta.objects.get(id=request.session.get('Huerta'))
            pedido = Pedido.objects.get(id=request.session.get('Pedido'))
            #Se obtienen los datos y se crea el miembro
            OrdenCorte.objects.create(
                fechaOrden=date.today(),
                numeroOrden=pedido.numeroPedido,
                idHuerta=huerta,
                tipoCorte=request.session.get('Corte'),
                idPedido=pedido
            )
            pedido.estatusPedido='P_O'
            pedido.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Orden de corte registrada correctamente."
        except Exception as e:
            print(e)
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('co')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def courtOrderModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):

        try:
            #Se obtienen los datos y se modifican

            calidad = Calidad.objects.get(id=request.session.get('Calidad'))
            calidad.numCalidad = request.session.get('Numero')
            calidad.descripcionCalidad = request.session.get('Descripcion')
            calidad.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('q')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def viewOrder(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B', 'I_C', 'E_R', 'E_A'):
        pedidos = Pedido.objects.all()
        p = Pedido.objects.filter(~Q(estatusPedido='P_C'))
        npedidosp = p.count()
        p = Pedido.objects.filter(estatusPedido='P_C')
        npedidost = p.count()
        pcc = PedidoCalibreCalidad.objects.all()

        return render(request, 'user_enc_bit/viewOrder.html', {
            'pedidos': pedidos,
            'pcc': pcc,
            'npedidosp': npedidosp,
            'npedidost': npedidost
        })
    else: 
        return render(request, 'denied.html')

@login_required
def incident(request):
    form = AddIncident()
    incidencias = Incidencia.objects.filter(idTrabajador=request.user.trabajador.id)
    nincidencias = incidencias.count()
    #si se envia un formulario
    if request.method == 'POST':
        if request.POST['Id']=='eliminar':
            try:
                request.session['Incidencia'] = request.POST['Incidencia']
                url = reverse('id')
                return redirect(url)
            except Exception as e:
                request.session['Operacion'] = 0
                request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                url = reverse('i')
                return redirect(url)
        elif request.POST['Id']=='agregar':
            try:
                request.session['Descripcion'] = request.POST['Descripcion']
                request.session['Fecha'] = request.POST['Fecha']
                request.session['Tema'] = request.POST['Tema']
                url = reverse('ir')
                return redirect(url)
            except Exception as e:
                request.session['Operacion'] = 0
                request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                url = reverse('i')
                return redirect(url)
    else:
        if request.session.get('Operacion')==1:
            request.session['Operacion'] = -1
            return render(request, 'incident.html', {
                'form':form,
                "mensaje": request.session['Mensaje'],
                'incidencias': incidencias,
                'nincidencias': nincidencias
            })
        elif request.session.get('Operacion')==0:
            request.session['Operacion'] = -1
            return render(request, 'incident.html', {
                'form':form,
                "error": request.session['Error'],
                'incidencias': incidencias,
                'nincidencias': nincidencias
            })
        else:
            return render(request, 'incident.html', {
                'form':form,
                'incidencias': incidencias,
                'nincidencias': nincidencias
            })

@login_required
def incidentDelete(request):
    try:
        #Se obtiene el miembro y elimina
        incidencia = request.session.get('Incidencia')
        i = Incidencia.objects.get(id=incidencia)
        i.delete()
        #Se guarda en memoria la operacion exitosa y redirige a la url de origen
        
        request.session['Operacion'] = 1
        request.session['Mensaje'] = "Reporte eliminado correctamente."
    except Exception as e:
        request.session['Operacion'] = 0
        request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
        
    url = reverse('i')
    return redirect(url)

@login_required
def incidentRegister(request):
    try:
        #Se obtienen los datos y se crea el reporte
        Incidencia.objects.create(
            descripcionIncidencia=request.session.get('Descripcion'),
            idTrabajador = request.user.trabajador,
            fechaIncidencia=request.session.get('Fecha'),
            temaIncidencia=request.session.get('Tema')
        )
        #Se guarda en memoria la operacion exitosa y redirige a la url de origen

        request.session['Operacion'] = 1
        request.session['Mensaje'] = "Se mando el reporte de la incidencia correctamente."
    except Exception as e:
        request.session['Operacion'] = 0
        request.session['Error'] = "No se pudo enviar el reporte, intente de nuevo."
        
    url = reverse('i')
    return redirect(url)

def safetyStatus(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_I'):
        huertas = Huerta.objects.all()
        nhuertas = huertas.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='modificar':
                try:
                    request.session['Huerta'] = request.POST['Huerta']
                    request.session['Clave'] = request.POST['Clave']
                    request.session['Inocuidad'] = request.POST['Inocuidad']
                    url = reverse('ssm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('ss')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ino/safetyStatusOrchard.html', {
                    "mensaje": request.session['Mensaje'],
                    'huertas': huertas,
                    'nhuertas': nhuertas
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_ino/safetyStatusOrchard.html', {
                    "error": request.session['Error'],
                    'huertas': huertas,
                    'nhuertas': nhuertas
                })
            else:
                return render(request, 'user_enc_ino/safetyStatusOrchard.html', {
                    'huertas': huertas,
                    'nhuertas': nhuertas
                })
    else: 
        return render(request, 'denied.html')

@login_required
def safetyStatusModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_I'):
        try:
            #Se obtienen los datos y se modifican
            huerta = Huerta.objects.get(id=request.session.get('Huerta'))
            huerta.claveSagarpaHuerta=request.session.get('Clave')
            huerta.estatusInocuidadHuerta=request.session.get('Inocuidad')
            huerta.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('ss')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

def finishTrip(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_R'):
        vc = ViajeCorte.objects.filter(idOrdenCorte__idPedido__estatusPedido='P_V')
        nvc = vc.count()
        pcc = PedidoCalibreCalidad.objects.all()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='agregar':
                try:
                    request.session['Viaje'] = request.POST['Viaje']
                    url = reverse('ftr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('ft')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_rec/finishTrip.html', {
                    "mensaje": request.session['Mensaje'],
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_enc_rec/finishTrip.html', {
                    "error": request.session['Error'],
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
            else:
                return render(request, 'user_enc_rec/finishTrip.html', {
                    'nvc': nvc,
                    'vc': vc,
                    'pcc': pcc
                })
    else: 
        return render(request, 'denied.html')

@login_required
def finishTripRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_R'):
        try:
            #Se obtienen los datos y se crea el miembro
            viaje = ViajeCorte.objects.get(id=request.session.get('Viaje'))
            cuadrilla = Cuadrilla.objects.get(id=viaje.idCuadrilla.id)
            orden = OrdenCorte.objects.get(id=viaje.idOrdenCorte.id)
            pedido = Pedido.objects.get(id=orden.idPedido.id)
            pedido.estatusPedido = 'P_C'
            cuadrilla.estatusCuadrilla = 'C_L'
            pedido.save()
            cuadrilla.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "El viaje ha sido completado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('ft')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def viewOrchard(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_T'):
        form = AddOrchard()
        huertas = Huerta.objects.all()
        nhuertas = huertas.count()

        return render(request, 'user_enc_trans/viewOrchard.html', {
            'form':form,
            'huertas': huertas,
            'nhuertas': nhuertas,
        })
    else: 
        return render(request, 'denied.html')
def myTrips(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('C_T','J_C'):

        viajes = ViajeCorte.objects.filter(Q(idCamionTransporte__idChoferTransporte=request.user.trabajador.id) | Q(idCuadrilla__idJefeCuadrilla=request.user.trabajador.id) | Q(idCamionSecundarioTransporte__idChoferTransporte=request.user.trabajador.id))
        viajes.order_by('-fechaViaje')
        nv = viajes.count()
        pcc = PedidoCalibreCalidad.objects.all()

        return render(request, 'user_con_trans/myTrips.html', {
            'viajes': viajes,
            'nv': nv,
            'pcc': pcc
        })
    else: 
        return render(request, 'denied.html')

@login_required
def viewSquad(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('C_T'):
        cuadrillas = Cuadrilla.objects.all()
        ncuadrillas = Cuadrilla.objects.all().count()
        miembros = MiembroCuadrilla.objects.all()

        return render(request, 'user_con_trans/viewSquad.html', {
            'ncuadrillas':ncuadrillas,
            'cuadrillas': cuadrillas,
            'miembros':miembros
        })
    else: 
        return render(request, 'denied.html')

@login_required
def mySquad(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        form = AddSquad() 
        cuadrillas = Cuadrilla.objects.filter(Q(idGerenteCuadrilla=request.user.trabajador.id) | Q(idJefeCuadrilla=request.user.trabajador.id))
        ncuadrillas = cuadrillas.count()
        numJefGer = Trabajador.objects.filter(Q(rol__nomenclaturaRol__exact='J_C')).count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    url = reverse('msd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('ms')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['Jefe'] = request.POST['Jefe']
                    request.session['Ubicacion'] = request.POST['Ubicacion']
                    request.session['Estatus'] = request.POST['Estatus']
                    url = reverse('msr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('ms')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Cuadrilla'] = request.POST['Cuadrilla']
                    url = reverse('msm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('ms')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_ger_cuad/mySquads.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    "mensaje": request.session['Mensaje'],
                    'cuadrillas': cuadrillas
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_ger_cuad/mySquads.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    "error": request.session['Error'],
                    'cuadrillas': cuadrillas
                })
            else:
                return render(request, 'user_ger_cuad/mySquads.html', {
                    'form':form,
                    'numJefGer':numJefGer,
                    'ncuadrillas':ncuadrillas,
                    'cuadrillas': cuadrillas
                })
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            #Se obtiene la cuadrilla y elimina
            cuadrilla = request.session.get('Cuadrilla')
            c = Cuadrilla.objects.get(id=cuadrilla)
            c.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cuadrilla eliminada correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('ms')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            #Se obtienen los datos y se crea la cuadrilla
            gerente=Trabajador.objects.get(id=request.user.trabajador.id)
            jefe=Trabajador.objects.get(id=request.session.get('Jefe'))
            Cuadrilla.objects.create(
                nombreCuadrilla=request.session.get('Nombre'),
                ubicacionCuadrilla=request.session.get('Ubicacion'),
                estatusCuadrilla=request.session.get('Estatus'),
                idJefeCuadrilla= jefe,
                idGerenteCuadrilla=gerente
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Cuadrilla registrada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('ms')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
        formm = AddSquadMember()
        miembros = MiembroCuadrilla.objects.filter(idCuadrilla=cuadrilla.id)
        trabajadores = Trabajador.objects.filter(Q(rol__nomenclaturaRol__exact='G_C') | Q(rol__nomenclaturaRol__exact='J_C'))
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Miembro'] = request.POST['Miembro']
                    url = reverse('msmd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('msm')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['noImss'] = request.POST['noImss']
                    url = reverse('msmr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('msm')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['Ubicacion'] = request.POST['Ubicacion']
                    request.session['Estatus'] = request.POST['Estatus']
                    if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
                        request.session['Jefe'] = request.POST['Jefe']
                    url = reverse('msms')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('msm')
                    return redirect(url)
            elif request.POST['Id']=='editar':
                try:
                    request.session['Miembro'] = request.POST['Miembro']
                    request.session['Nombre'] = request.POST['Nombre']
                    request.session['AP'] = request.POST['AP']
                    request.session['AM'] = request.POST['AM']
                    request.session['noImss'] = request.POST['noImss']
                    url = reverse('msmm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo guardar los datos, intente de nuevo."
                    url = reverse('msm')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_ger_cuad/mySquadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    "mensaje": request.session['Mensaje'],
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_ger_cuad/mySquadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    "error": request.session['Error'],
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
            else:
                return render(request, 'user_ger_cuad/mySquadModify.html', {
                    'formm':formm,
                    'cuadrilla':cuadrilla,
                    'trabajadores' : trabajadores,
                    'miembros': miembros
                })
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadMemberSave(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            #Se obtienen los datos y se modifican
            if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
                jefe=Trabajador.objects.get(id=request.session.get('Jefe'))

            cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
            cuadrilla.nombreCuadrilla=request.session.get('Nombre')
            cuadrilla.estatusCuadrilla=request.session.get('Estatus')
            cuadrilla.ubicacionCuadrilla=request.session.get('Ubicacion')
            if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
                cuadrilla.idJefeCuadrilla=jefe
            cuadrilla.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron los datos correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('msm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadMemberDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            #Se obtiene el miembro y elimina
            miembro = request.session.get('Miembro')
            m = MiembroCuadrilla.objects.get(id=miembro)
            m.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Miembro eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('msm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadMemberModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            #Se obtienen los datos y se modifican
            miembro = MiembroCuadrilla.objects.get(id=request.session.get('Miembro'))
            miembro.nombre=request.session.get('Nombre')
            miembro.apellidoP=request.session.get('AP')
            miembro.apellidoM=request.session.get('AM')
            miembro.noImss = request.session.get('noImss')
            miembro.save()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('msm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def mySquadMemberRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C', 'J_C'):
        try:
            cuadrilla = Cuadrilla.objects.get(id=request.session.get('Cuadrilla'))
            #Se obtienen los datos y se crea el miembro
            MiembroCuadrilla.objects.create(
                nombre=request.session.get('Nombre'),
                apellidoP=request.session.get('AP'),
                apellidoM=request.session.get('AM'),
                noImss = request.session.get('noImss'),
                idCuadrilla=cuadrilla
            )
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Miembro registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('msm')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadLeader(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
        #Consultas necesarias para mostrar en plantilla
        form = AddWorker()
        jefes = Trabajador.objects.filter(rol__nomenclaturaRol='J_C')
        njefes = jefes.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Trabajador'] = request.POST['Trabajador']
                    url = reverse('sld')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('sl')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nombres'] = request.POST['Nombre']
                    request.session['Apellidos'] = request.POST['AP']+' '+request.POST['AM']
                    request.session['Telefono'] = request.POST['Telefono']
                    request.session['Correo'] = request.POST['Correo']
                    url = reverse('slr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('sl')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Trabajador'] = request.POST['Trabajador']
                    request.session['Nusuario'] = request.POST['Nusuario']
                    request.session['Nombres'] = request.POST['Nombre']
                    request.session['Apellidos'] = request.POST['Apellidos']
                    request.session['Telefono'] = request.POST['Telefono']
                    request.session['Correo'] = request.POST['Correo']
                    url = reverse('slm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('sl')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_ger_cuad/squadLeader.html", {
                    'form':form,
                    'jefes': jefes,
                    'njefes': njefes,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_ger_cuad/squadLeader.html", {
                    'form':form,
                    'jefes': jefes,
                    'njefes': njefes,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_ger_cuad/squadLeader.html", {
                    'form':form,
                    'jefes': jefes,
                    'njefes': njefes,
                })
    else: 
        return render(request, 'denied.html')

@login_required
def squadLeaderDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
        try:
            #Se obtiene el trabajador y elimina
            trabajador = request.session.get('Trabajador')
            t = Trabajador.objects.get(id=trabajador)
            u = User.objects.get(id = t.usuario.id)
            u.delete() 
            t.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Trabajador eliminado correctamente."
        except Exception as e:

            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('sl')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadLeaderRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
        try:
            nombres=request.session.get('Nombres')
            apellidos=request.session.get('Apellidos')
            palabras = nombres.split() + apellidos.split()
            iniciales = []
            for palabra in palabras:
                iniciales.append(palabra[0].upper())
            nusuario = ''.join(iniciales)
            usuario = User.objects.create(
                username='Agro-'+nusuario,
                first_name=request.session.get('Nombres'),
                last_name=request.session.get('Apellidos'),
                password= make_password(request.session.get('Telefono')),
                email = request.session.get('Correo')
            )

            r = RolTrabajador.objects.get(nomenclaturaRol='J_C')
            Trabajador.objects.create(
                telefono=request.session.get('Telefono'),
                rol=r,
                usuario=usuario
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Jefe de cuadrilla registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('sl')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def squadLeaderModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('G_C'):
        try:
            #Se obtienen los datos y se modifican
            trabajador = Trabajador.objects.get(id=request.session.get('Trabajador'))
            usuario = User.objects.get(id=trabajador.usuario.id)
            usuario.username=request.session.get('Nusuario')
            usuario.first_name =request.session.get('Nombres')
            usuario.last_name=request.session.get('Apellidos')
            usuario.email=request.session.get('Correo')
            usuario.save()
            trabajador.telefono=request.session.get('Telefono')
            trabajador.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('sl')
        return redirect(url)
    else: 
        return render(request, 'denied.html')


@login_required
def solveIncident(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        incidencias = Incidencia.objects.all()
        nincidencias = incidencias.count()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Incidencia'] = request.POST['Incidencia']
                    url = reverse('sid')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('si')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Incidencia'] = request.POST['Incidencia']
                    request.session['Solucion'] = request.POST['Solucion']
                    url = reverse('sir')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo enviar el cometido, intente de nuevo."
                    url = reverse('si')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, 'user_dir_gral/resolveIncident.html', {
                    "mensaje": request.session['Mensaje'],
                    'incidencias': incidencias,
                    'nincidencias': nincidencias
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, 'user_dir_gral/resolveIncident.html', {
                    "error": request.session['Error'],
                    'incidencias': incidencias,
                    'nincidencias': nincidencias
                })
            else:
                return render(request, 'user_dir_gral/resolveIncident.html', {
                    'incidencias': incidencias,
                    'nincidencias': nincidencias
                })
    else:
        return render(request, 'denied.html')

@login_required
def solveIncidentDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        try:
            #Se obtiene el miembro y elimina
            incidencia = request.session.get('Incidencia')
            i = Incidencia.objects.get(id=incidencia)
            i.delete()
            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Reporte eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('si')
        return redirect(url)
    else:
        return render(request, 'denied.html')

@login_required
def solveIncidentRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        try:
            #Se obtienen los datos y se crea el reporte
            i = Incidencia.objects.get(id=request.session.get('Incidencia'))
            i.estatusIncidencia='I_R'
            i.descripcionSolucion= request.session.get('Solucion')
            i.save()

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se mando la solución de la incidencia correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la acción, intente de nuevo."
            
        url = reverse('si')
        return redirect(url)
    else:
        return render(request, 'denied.html')

@login_required
def authorizedPrice(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        #Consultas necesarias para mostrar en plantilla
        estados = (
            ('AGS', 'Aguascalientes'),
            ('BC', 'Baja California'),
            ('BCS', 'Baja California Sur'),
            ('CAM', 'Campeche'),
            ('CHIS', 'Chiapas'),
            ('CHIH', 'Chihuahua'),
            ('CDMX', 'Ciudad de México'),
            ('COAH', 'Coahuila'),
            ('COL', 'Colima'),
            ('DGO', 'Durango'),
            ('GTO', 'Guanajuato'),
            ('GRO', 'Guerrero'),
            ('HGO', 'Hidalgo'),
            ('JAL', 'Jalisco'),
            ('MEX', 'México'),
            ('MIC', 'Michoacán'),
            ('MOR', 'Morelos'),
            ('NAY', 'Nayarit'),
            ('NL', 'Nuevo León'),
            ('OAX', 'Oaxaca'),
            ('PUE', 'Puebla'),
            ('QRO', 'Querétaro'),
            ('QR', 'Quintana Roo'),
            ('SLP', 'San Luis Potosí'),
            ('SIN', 'Sinaloa'),
            ('SON', 'Sonora'),
            ('TAB', 'Tabasco'),
            ('TAMPS', 'Tamaulipas'),
            ('TLAX', 'Tlaxcala'),
            ('VER', 'Veracruz'),
            ('YUC', 'Yucatán'),
            ('ZAC', 'Zacatecas'),
        )
        form = AddPrice()
        precios = PrecioAutorizado.objects.all()
        pcr = PrecioAutorizado.objects.values_list('estadoAplica', flat=True).distinct()
        psr = [estado for estado in estados if estado[0] not in pcr]
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Precio'] = request.POST['Precio']
                    url = reverse('apd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('ap')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Fijo'] = request.POST['Fijo']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    request.session['Vigencia'] = request.POST['Vigencia']
                    request.session['Estado'] = request.POST['Estado']
                    url = reverse('apr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('ap')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Precio'] = request.POST['Precio']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    request.session['Actual'] = request.POST['Actual']
                    request.session['Vigencia'] = request.POST['Vigencia']
                    request.session['Estado'] = request.POST['Estado']
                    url = reverse('apm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('ap')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/authorizedPrice.html", {
                    'form':form,
                    'estados': estados,
                    'precios': precios,
                    'psr': psr,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/authorizedPrice.html", {
                    'form':form,
                    'estados': estados,
                    'precios': precios,
                    'psr': psr,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_enc_bit/authorizedPrice.html", {
                    'form':form,
                    'estados': estados,
                    'precios': precios,
                    'psr': psr,
                })
    else: 
        return render(request, 'denied.html')

@login_required
def authorizedPriceDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtiene el trabajador y elimina
            precio = request.session.get('Precio')
            p = PrecioAutorizado.objects.get(id=precio)
            p.delete()

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Precio eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('ap')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def authorizedPriceRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            PrecioAutorizado.objects.create(
                precioFijo=request.session.get('Fijo'),
                descripcion=request.session.get('Descripcion'),
                precioActual=request.session.get('Fijo'),
                vigencia=request.session.get('Vigencia'),
                estadoAplica=request.session.get('Estado'),
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Precio registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('ap')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def authorizedPriceModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtienen los datos y se modifican
            precio = PrecioAutorizado.objects.get(id=request.session.get('Precio'))
            precio.descripcion =request.session.get('Descripcion')
            pa = request.session.get('Actual')
            pa = pa.replace(',', '.')
            precio.precioActual=float(pa)
            precio.vigencia=request.session.get('Vigencia')
            precio.estadoAplica=request.session.get('Estado')
            precio.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            print(e)
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('ap')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def fruit(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_A'):
        #Consultas necesarias para mostrar en plantilla
        form = AddFruit()
        huertas = Huerta.objects.exclude(frutahuerta__isnull=False)
        fruta = FrutaHuerta.objects.all()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Fruta'] = request.POST['Fruta']
                    url = reverse('fd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('f')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Huerta'] = request.POST['Huerta']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    request.session['Precio'] = request.POST['Precio']
                    url = reverse('fr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('f')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Fruta'] = request.POST['Fruta']
                    request.session['Huerta'] = request.POST['Huerta']
                    request.session['Descripcion'] = request.POST['Descripcion']
                    request.session['Precio'] = request.POST['Precio']
                    url = reverse('fm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('f')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_enc_aco/fruit.html", {
                    'form':form,
                    'fruta': fruta,
                    'huertas': huertas,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_enc_aco/fruit.html", {
                    'form':form,
                    'fruta': fruta,
                    'huertas': huertas,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_enc_aco/fruit.html", {
                    'form':form,
                    'huertas': huertas,
                    'fruta': fruta,
                })
    else: 
        return render(request, 'denied.html')

@login_required
def fruitDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_A'):
        try:
            #Se obtiene el trabajador y elimina
            fruta = request.session.get('Fruta')
            f = FrutaHuerta.objects.get(id=fruta)
            f.delete()

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Fruta eliminada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('f')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def fruitRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_A'):
        try:
            huerta = Huerta.objects.get(id=request.session.get('Huerta'))
            FrutaHuerta.objects.create(
                idHuerta=huerta,
                descripcionFruta=request.session.get('Descripcion'),
                precioFruta=request.session.get('Precio'),
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Fruta registrada correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('f')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def fruitModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_A'):
        try:
            #Se obtienen los datos y se modifican
            p = request.session.get('Precio')
            p = p.replace(',', '.')
            fruta = FrutaHuerta.objects.get(id=request.session.get('Fruta'))
            huerta = Huerta.objects.get(id=request.session.get('Huerta'))
            fruta.huerta = huerta
            fruta.descripcionFruta = request.session.get('Descripcion')
            fruta.precioFruta = float(p)
            fruta.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            print(e)
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('f')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def rol(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        #Consultas necesarias para mostrar en plantilla
        form = AddRol()
        roles = RolTrabajador.objects.all()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Rol'] = request.POST['Rol']
                    url = reverse('rd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('r')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Nomenclatura'] = request.POST['Nomenclatura']
                    request.session['Nombre'] = request.POST['Nombre']
                    url = reverse('rr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('r')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    request.session['Rol'] = request.POST['Rol']
                    request.session['Nomenclatura'] = request.POST['Nomenclatura']
                    request.session['Nombre'] = request.POST['Nombre']
                    url = reverse('rm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('r')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/rol.html", {
                    'form':form,
                    'roles': roles,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_enc_bit/rol.html", {
                    'form':form,
                    'roles': roles,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_enc_bit/rol.html", {
                    'form':form,
                    'roles': roles,
                })
    else: 
        return render(request, 'denied.html')

@login_required
def rolDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtiene el trabajador y elimina
            r = RolTrabajador.objects.get(id=request.session.get('Rol'))
            r.delete()

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Rol eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('r')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def rolRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            RolTrabajador.objects.create(
                nomenclaturaRol=request.session.get('Nomenclatura'),
                nombreRol=request.session.get('Nombre'),
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Rol registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('r')
        return redirect(url)
    else:
        return render(request, 'denied.html')

@login_required
def rolModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('E_B'):
        try:
            #Se obtienen los datos y se modifican
            rol = RolTrabajador.objects.get(id=request.session.get('Rol'))
            rol.nombreRol = request.session.get('Nombre')
            rol.nomenclaturaRol = request.session.get('Nomenclatura')
            rol.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('r')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def report(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('J_C'):
        #Consultas necesarias para mostrar en plantilla
        form = AddReport(request=request)
        reportes = ReporteCorte.objects.all()
        viajes = ViajeCorte.objects.filter(idCuadrilla__idJefeCuadrilla=request.user.trabajador.id).exclude(reportecorte__idViaje__isnull=False)
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='eliminar':
                try:
                    request.session['Reporte'] = request.POST['Reporte']
                    url = reverse('rpd')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
                    url = reverse('rp')
                    return redirect(url)
            elif request.POST['Id']=='agregar':
                try:
                    request.session['Viaje'] = request.POST['Viaje']
                    request.session['Observaciones'] = request.POST['Observaciones']
                    request.session['Cajas'] = request.POST['Cajas']
                    url = reverse('rpr')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
                    url = reverse('rp')
                    return redirect(url)
            elif request.POST['Id']=='modificar':
                try:
                    url = reverse('rpm')
                    return redirect(url)
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
                    url = reverse('rp')
                    return redirect(url)
        else:
            if request.session.get('Operacion')==1:
                request.session['Operacion'] = -1
                return render(request, "user_jefe_cuad/report.html", {
                    'form':form,
                    'reportes': reportes,
                    'viajes': viajes,
                    "mensaje": request.session['Mensaje']
                })
            elif request.session.get('Operacion')==0:
                request.session['Operacion'] = -1
                return render(request, "user_jefe_cuad/report.html", {
                    'form':form,
                    'reportes': reportes,
                    'viajes': viajes,
                    "error": request.session['Error']
                })
            else:
                return render(request, "user_jefe_cuad/report.html", {
                    'form':form,
                    'reportes': reportes,
                    'viajes': viajes,
                })
    else: 
        return render(request, 'denied.html')

@login_required
def reportDelete(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('J_C'):
        try:
            #Se obtiene el trabajador y elimina
            r = ReporteCorte.objects.get(id=request.session.get('Reporte'))
            r.delete()

            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Reporte eliminado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar la eliminación, intente de nuevo."
            
        url = reverse('rp')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def reportRegister(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('J_C'):
        try:
            viaje = ViajeCorte.objects.get(id = request.session.get('Viaje'))
            ReporteCorte.objects.create(
                fecha = date.today(),
                idViaje=viaje,
                observacionesReporte=request.session.get('Observaciones'),
                cajasCortadas=request.session.get('Cajas'),
            )
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Reporte registrado correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo realizar el registro, intente de nuevo."
            
        url = reverse('rp')
        return redirect(url)
    else:
        return render(request, 'denied.html')

@login_required
def reportModify(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('J_C'):
        try:
            #Se obtienen los datos y se modifican
            rol = RolTrabajador.objects.get(id=request.session.get('Rol'))
            rol.nombreRol = request.session.get('Nombre')
            rol.nomenclaturaRol = request.session.get('Nomenclatura')
            rol.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('r')
        return redirect(url)
    else: 
        return render(request, 'denied.html')

@login_required
def cutLog(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        form = SearchTrip()
        #si se envia un formulario
        if request.method == 'POST':
            if request.POST['Id']=='buscar':
                try:
                    i = request.POST['Inicio']
                    f = request.POST['Fin']
                    viajes = list(ViajeCorte.objects.filter(Q(fechaViaje__range=[i, f])))
                    return render(request, "user_dir_gral/cutLog.html", {
                        'form':form,
                        "viajes": viajes
                    })
                except Exception as e:
                    request.session['Operacion'] = 0
                    request.session['Error'] = "Ha ocurrido un error, intente de nuevo."
                    url = reverse('cl')
                    return redirect(url)
            elif request.POST['Id']=='generar':
                try:
                    filas = request.POST['Filas'].split(',')
                    viajes = ViajeCorte.objects.filter(id__in=filas)
                    resultados = viajes.aggregate(primera_fecha=Min('fechaViaje'), ultima_fecha=Max('fechaViaje'))
                    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
                    i = resultados['primera_fecha']
                    f = resultados['ultima_fecha']
                    i_texto = i.strftime('%d de %B de %Y')
                    f_texto = f.strftime('%d de %B de %Y')
                    ruta = os.path.join(os.getcwd(), 'Agronomunnity', 'static', 'temp')
                    os.makedirs(ruta, exist_ok=True)
                    ruta = os.path.join(ruta, 'bitacora.xlsx')
                    wb = Workbook()
                    #titulo de la hoja
                    hoja = wb.active
                    hoja.title = "Bitacora"
                    #Logos
                    hoja.merge_cells('A1:F5')
                    hoja['A1'] = 'FRUTIVAL SA DE CV'
                    fuente = Font(name='Century Gothic', size=20, bold=True)
                    hoja['A1'].font = fuente
                    hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    hoja.merge_cells('G1:G5')
                    rimagen = os.path.join(os.getcwd(), 'Agronomunnity', 'static', 'assets', 'img')
                    os.makedirs(rimagen, exist_ok=True)
                    rimagen = os.path.join(rimagen, 'frutival.jpg')
                    imagen = Image(rimagen)
                    imagen.anchor = 'G1'
                    imagen.width = 150
                    imagen.height = 100
                    hoja.add_image(imagen)
                    hoja.merge_cells('H1:O5')
                    texto = 'Bitacora de reportes del {} al {}'.format(i_texto, f_texto)
                    hoja['H1'] = texto
                    fuente = Font(name='Arial', size=11)
                    hoja['H1'].font = fuente
                    hoja['H1'].alignment = Alignment(horizontal='center', vertical='center')
                    #tabla de viajes
                    campos = ['FECHA', 'HUERTA', 'PRODUCTOR', 'TIPO DE CORTE', 'PUNTO DE REUNION',
                              'CAMION','CONDUCTOR', 'CUADRILLA', 'CANDADO', 'KILO',
                              'PALETS', 'MERCADO','DESTINO', 'CLIENTE', 'OBSERVACIONES']
                    for col, campo in enumerate(campos, start=1):
                        cell = hoja.cell(row=7, column=col)
                        cell.value = campo
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(fill_type="solid", fgColor="4CAF50")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        hoja.cell(row=7, column=col).alignment = Alignment(horizontal='center', vertical='center')

                    # Insertar los datos de la consulta
                    fila_inicio_datos = 8
                    for i, viaje in enumerate(viajes, start=fila_inicio_datos):
                        hoja.cell(row=i, column=1, value=viaje.fechaViaje)
                        hoja.cell(row=i, column=2, value=viaje.idOrdenCorte.idHuerta.nombreHuerta)
                        hoja.cell(row=i, column=3, value=f"{viaje.idOrdenCorte.idHuerta.idProductor.nombre} {viaje.idOrdenCorte.idHuerta.idProductor.apellidoP} {viaje.idOrdenCorte.idHuerta.idProductor.apellidoM}")
                        hoja.cell(row=i, column=4, value=viaje.idOrdenCorte.get_tipoCorte_display())
                        hoja.cell(row=i, column=5, value=viaje.puntoReunion)
                        hoja.cell(row=i, column=6, value=viaje.idCamionTransporte.modeloTransporte)
                        hoja.cell(row=i, column=7, value=f"{viaje.idCamionTransporte.idChoferTransporte.usuario.first_name} {viaje.idCamionTransporte.idChoferTransporte.usuario.last_name}")
                        hoja.cell(row=i, column=8, value=viaje.idCuadrilla.nombreCuadrilla)
                        hoja.cell(row=i, column=9, value=viaje.idCamionTransporte.candadoTransporte)
                        hoja.cell(row=i, column=10, value=viaje.idOrdenCorte.idPedido.totalKilosPedido)
                        hoja.cell(row=i, column=11, value=viaje.idOrdenCorte.idPedido.totalPalletsPedido)
                        hoja.cell(row=i, column=12, value=viaje.idOrdenCorte.idPedido.get_mercadoPedido_display())
                        hoja.cell(row=i, column=13, value=viaje.idOrdenCorte.idPedido.destinoPedido)
                        hoja.cell(row=i, column=14, value=f"{viaje.idOrdenCorte.idPedido.idCliente.nombreCliente} {viaje.idOrdenCorte.idPedido.idCliente.apellidoPCliente} {viaje.idOrdenCorte.idPedido.idCliente.apellidoMCliente}")
                        hoja.cell(row=i, column=15, value=viaje.idOrdenCorte.idPedido.observacionPedido)
                        fila_inicio_datos = i
                    # Ajustar el ancho de las columnas al tamaño del título
                    for col in range(1, len(campos) + 1):
                        column_letter = get_column_letter(col)
                        hoja.column_dimensions[column_letter].auto_size = True
                    
                    fila_inicio_datos +=3
                    cf = get_column_letter(3)
                    ff = get_column_letter(6)
                    hoja.merge_cells(f'{cf}{fila_inicio_datos}:{ff}{fila_inicio_datos}')
                    texto = 'PRECIOS AUTORIZADOS'
                    celda = hoja[f'{cf}{fila_inicio_datos}']
                    celda.value = texto
                    celda.font = Font(name='Arial', size=11)
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    celda.fill = PatternFill(fill_type="solid", fgColor="4CAF50")
                    celda.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    cf = get_column_letter(9)
                    ff = get_column_letter(11)
                    hoja.merge_cells(f'{cf}{fila_inicio_datos}:{ff}{fila_inicio_datos}')
                    texto = 'FRUTA DE HUERTAS'
                    celda = hoja[f'{cf}{fila_inicio_datos}']
                    celda.value = texto
                    celda.font = Font(name='Arial', size=11)
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    celda.fill = PatternFill(fill_type="solid", fgColor="4CAF50")
                    celda.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    fila_inicio_datos +=1

                     #tabla de precios autorizados
                    campos = ['ESTADO', 'DESCRIPCION', 'PRECIO', 'VIGENCIA']
                    for col, campo in enumerate(campos, start=3):
                        cell = hoja.cell(row=fila_inicio_datos, column=col)
                        cell.value = campo
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(fill_type="solid", fgColor="4CAF50")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        hoja.cell(row=fila_inicio_datos, column=col).alignment = Alignment(horizontal='center', vertical='center')

                    #tabal de fruta huerta
                    campos = ['HUERTA', 'FRUTA', 'PRECIO']
                    for col, campo in enumerate(campos, start=9):
                        cell = hoja.cell(row=fila_inicio_datos, column=col)
                        cell.value = campo
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(fill_type="solid", fgColor="4CAF50")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        hoja.cell(row=fila_inicio_datos, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    fila_inicio_datos +=1
                    precios = PrecioAutorizado.objects.all()
                    x=0
                    for i, precio in enumerate(precios, start=fila_inicio_datos):
                        hoja.cell(row=i, column=3, value=precio.get_estadoAplica_display())
                        hoja.cell(row=i, column=4, value=precio.descripcion)
                        hoja.cell(row=i, column=5, value=precio.precioActual)
                        hoja.cell(row=i, column=6, value=precio.vigencia)
                        x+=1
                    frutas = FrutaHuerta.objects.all()
                    y=0
                    for i, fruta in enumerate(frutas, start=fila_inicio_datos):
                        hoja.cell(row=i, column=9, value=fruta.idHuerta.nombreHuerta)
                        hoja.cell(row=i, column=10, value=fruta.descripcionFruta)
                        hoja.cell(row=i, column=11, value=fruta.precioFruta)
                        y+=1
                    fila_inicio_datos += x if x > y else y
                    
                    fila_inicio_datos += 2
                    cf = get_column_letter(3)
                    ff = get_column_letter(6)
                    hoja.merge_cells(f'{cf}{fila_inicio_datos}:{ff}{fila_inicio_datos}')
                    ea= Trabajador.objects.get(rol_id=12)
                    texto = 'Precios Autorizados por {} {}.'.format(ea.usuario.first_name, ea.usuario.last_name)
                    celda = hoja[f'{cf}{fila_inicio_datos}']
                    celda.value = texto
                    celda.font = Font(name='Arial', size=9)
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    celda.fill = PatternFill(fill_type="solid", fgColor="FFAE9D")
                    celda.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    cf = get_column_letter(8)
                    ff = get_column_letter(11)
                    hoja.merge_cells(f'{cf}{fila_inicio_datos}:{ff}{fila_inicio_datos}')
                    ea= request.user.trabajador
                    texto = 'Bitacora generada por {} {}.'.format(ea.usuario.first_name, ea.usuario.last_name)
                    celda = hoja[f'{cf}{fila_inicio_datos}']
                    celda.value = texto
                    celda.font = Font(name='Arial', size=9)
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    celda.fill = PatternFill(fill_type="solid", fgColor="FFAE9D")
                    celda.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    fila_inicio_datos +=1

                    wb.save(ruta)
                    wb.close()
                    return render(request, "user_dir_gral/cutLog.html", {
                        'form':form,
                        'listo':True,
                        'ruta': ruta
                    })
                except Exception as e:
                    print(e)
                    request.session['Operacion'] = 0
                    request.session['Error'] = "Ha ocurrido un error, intente de nuevo."
                    url = reverse('cl')
                    return redirect(url)
        else:
            
            return render(request, "user_dir_gral/cutLog.html", {
                'form':form,
            })
    else: 
        return render(request, 'denied.html')

@login_required
def cutLogSearch(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        try:
            i = request.session.get('Inicio')
            f = request.session.get('Fin')
            viajes = list(ViajeCorte.objects.filter(Q(fechaViaje__range=[i, f])))

            request.session['Operacion'] = 1
            request.session['Viajes'] = viajes
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "Ha ocurrido un error, intente de nuevo."
            
        url = reverse('cl')
        return redirect(url)
    else:
        return render(request, 'denied.html')

@login_required
def cutLogGenerate(request):
    if request.user.trabajador.rol.nomenclaturaRol in ('D_G'):
        try:
            #Se obtienen los datos y se modifican
            rol = RolTrabajador.objects.get(id=request.session.get('Rol'))
            rol.nombreRol = request.session.get('Nombre')
            rol.nomenclaturaRol = request.session.get('Nomenclatura')
            rol.save()

            #Se guarda en memoria la operacion exitosa y redirige a la url de origen
            request.session['Operacion'] = 1
            request.session['Mensaje'] = "Se guardaron las modificaciones correctamente."
        except Exception as e:
            request.session['Operacion'] = 0
            request.session['Error'] = "No se pudo modificar los datos, intente de nuevo."
            
        url = reverse('r')
        return redirect(url)
    else: 
        return render(request, 'denied.html')
    

#pwa 
def offline(request):
    return render(request, 'offline.html')